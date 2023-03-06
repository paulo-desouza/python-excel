from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter
import datetime 
from operator import itemgetter
import pandas as pd
import pyexcel as p
import os



def get_file_names():
    """
    Returns
    -------
    All filenames of this python file's current directory, excluding itself. 
    
    """

    current_dir = os.getcwd()

    file_list = os.listdir(current_dir)

    file_list.remove("data_merger.py")

    return file_list


def set_border(ws, cell_range):
    thin = Side(border_style="medium", color="000000")
    for row in ws[cell_range]:
        for cell in row:
            cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)


def month_year_presenter(date):
    
    if "/" in date:
        strip1 = date.split("/")
    else:
        return "--"
    
    
    datetime1 = datetime.datetime(int(strip1[2]), int(strip1[0]), int(strip1[1]))  # y, m, d
    
    return datetime1.strftime('%B, %Y')


def days_to_months(dt_days):
    
    if dt_days == str(dt_days):
        return dt_days
    
    
    days = dt_days.days
    months = 0
    while days > 30:
        months += 1
        days -= 30
    return str(months)


def days_int(td):
    
    if td == str(td):
        return td
    else:
        int_days = int(td.days)
        return int_days
    

def time_check(td, time=75):
    """
    Parameters
    ----------
    td : deltatime
        Deltatime value.
    time : Integer
        Day limit that determines if we are on or off track. The default is 75.

    Returns
    -------
    """
    on_track = True
    int_days = 0
    int_months = 0
    
    if td == str(td):
        return td
    else:
        
        int_days = int(td.days)
        if int_days > time:
            on_track = False
        
        
        while int_days > 30:
            int_months += 1
            int_days -= 30
        
        if time == 0:
            if int_months == 0:
                return str(int_days) + " day(s)"
            else:
                return str(int_months) + " month(s) and " + str(int_days) + " day(s)"
        if on_track:
            if int_months == 0:
                return "On Track (" + str(int_days) + " day(s))" 
            else:
                return "On Track (" + str(int_months) + " month(s) and " + str(int_days) + " day(s))" 
        else:
            if int_months == 0:
                return "Off Track (" + str(int_days) + " day(s))" 
            else:
                return "Off Track (" + str(int_months) + " month(s) and " + str(int_days) + " day(s))" 
            
        
        


def date_subtraction(start, finish, format="days"):
    """
    Parameters
    ----------
    start : MM/DD/YYYY
    finish : MM/DD/YYYY
    
    Desc: Converts both dates into DateTime Objects and subtracts them from one another.
    
    Returns 
    -------
    Time period in between the 2 dates.

    """
    
    if "/" in start and "/" in finish:
    
        strip1 = start.split("/")
        strip2 = finish.split("/")
        
    else:
        
        return "--"
    
    
    datetime1 = datetime.datetime(int(strip1[2]), int(strip1[0]), int(strip1[1]))  # y, m, d
    datetime2 = datetime.datetime(int(strip2[2]), int(strip2[0]), int(strip2[1]))
    
    return datetime2 - datetime1
    



def time_from_present(start, present=datetime.datetime.now(), format="days"):
    """
    Parameters
    ----------
    start : MM/DD/YYYY
    
    Desc: Converts the date into DateTime Object and subtracts it from the present.
    
    Returns 
    -------
    Time period in between the date and the present.

    """
    
    if "/" in start:
    
        strip1 = start.split("/")

        
    else:
        
        return "--"
    
    
    datetime1 = datetime.datetime(int(strip1[2]), int(strip1[0]), int(strip1[1]))  # y, m, d
    
    return present - datetime1

def time_converter(date):
    
    if "/" in date:
        strip1 = date.split("/")
    
    datetime1 = datetime.datetime(int(strip1[2]), int(strip1[0]), int(strip1[1]))  # y, m, d

    return datetime1        
    
    
    

# CREATE A DICTIONARY WITH KEY AS CUSTOMER ID, AND VALUE AS LIST WITH CUSTOMER 
# INFORMATION.  FIRST ITEM HAS KEY "ID" AND VALUE AS LIST WITH ALL COLUMN IDs


datadump1_organized = {}
datadump2_organized = {}

# file conversion

file_list = get_file_names()


if 'Store Opening Timeline Report_from Franconnect.xls' in file_list:
    p.save_book_as(file_name='Store Opening Timeline Report_from Franconnect.xls',
                   dest_file_name='Store Opening Timeline Report_from Franconnect.xlsx')
    os.remove('Store Opening Timeline Report_from Franconnect.xls')
    
    
elif 'Store Opening Timeline Report_from Franconnect.xlsx' in file_list:
    pass


wb = load_workbook("Store Opening Timeline Report_from Franconnect.xlsx")
ws = wb.active

# DESCRIPTIONS OF CLIENT IDs HAVE TO MATCH
ws["A2"].value = "Franchise ID"


# COUNT THE ROWS

count = 0
for row in ws:
    if not all([cell.value is None for cell in row]):
        count += 1


# GET ALL CLIENTS IN THESE DICTS

for row in range(12, count + 1):
    for column in range(1, 36):
        char = get_column_letter(column)
        
        if char == 'A':
            datadump1_organized[ws[char + str(row)].value] = []
            
        else:
            datadump1_organized[ws['A' + str(row)].value] += [ws[char + str(row)].value]       


# file conversion




if 'Store Summary Dashboard Report_from Franconnect.xls' in file_list:
    p.save_book_as(file_name='Store Summary Dashboard Report_from Franconnect.xls',
                   dest_file_name='Store Summary Dashboard Report_from Franconnect.xlsx')
    os.remove('Store Summary Dashboard Report_from Franconnect.xls')
    
elif 'Store Opening Timeline Report_from Franconnect.xlsx' in file_list:
    pass


wb = load_workbook("Store Summary Dashboard Report_from Franconnect.xlsx")
ws = wb.active


# count the amount of rows with data
count = 0
for row in ws:
    if not all([cell.value is None for cell in row]):
        count += 1


for row in range(3, count + 1):
    for column in range(1, 8):
        char = get_column_letter(column)
        
        if char == 'A':
            datadump2_organized[ws[char + str(row)].value] = []
            
        else:
            datadump2_organized[ws['A' + str(row)].value] += [ws[char + str(row)].value]
    

    
# SORT THE DICTIONARIES:
    # IDEAL: KEEP THE ORDER OF THE SHEETS.
    # FOR EACH ITEM OF DIC1, LOOP THROUGH ALL OF DIC2. 
    # IF YOU FIND THE MATCH, MERGE IT
    # ELSE, ADD IT ELSEWHERE SO WE KNOW
    
    
filler_list = ["--", "--", "--", "--", "--", "--"]
joint_data = {}
datadump1_leftovers = datadump1_organized.copy()
datadump2_leftovers = datadump2_organized.copy()

for client1 in datadump1_organized:
    for client2 in datadump2_organized:
        if client1 in client2:
            
            client = client1
            
            joint_data[client] = datadump1_organized[client1]
            joint_data[client] += datadump2_organized[client2]
            
            del datadump1_leftovers[client1]
            del datadump2_leftovers[client2]
        
for client in datadump1_leftovers:
    
    joint_data[client] = datadump1_leftovers[client]
    joint_data[client] += filler_list
    

keys = list(joint_data)    

# FROM JOINT DATA,MAKE ONE SEPARATE LIST WITH THE DATA NEEDED FOR  
# EACH OF THE TABLES.





contruction_pipeline = []

# loop through the dictionary, and during each iteration, pull and append the data you need to the list.

AVG_SUM_CP1 = 0
AVG_SUM_CP2 = 0


AVG_COUNT_CP = 0



for c, client in enumerate(joint_data):
    if c == 0:
        pass
    
    else:
        
        if joint_data[client][1].strip() == '--':
            
            pass
        
        else:
        
            if joint_data[client][13].strip() == "--" :
                finance = "In Process / Self Funded"
            else:
                finance = "Funded"
            
            # Filtering by STATUS
            if joint_data[client][35] == "School Fit-Out":
            
                contruction_pipeline.append(
                    {   "Franchise ID" : client,
                        "Project Status" : joint_data[client][35],
                        "City, State" : joint_data[client][37] + ", " + joint_data[client][38],
                        "Architecturals (2.5 Months)" : time_check(date_subtraction(joint_data[client][9], joint_data[client][10])),
                        "Permitting (2.5 Months)" : time_check(date_subtraction(joint_data[client][10], joint_data[client][11])),
                        "Active Construction (5 months)" : time_check(time_from_present(joint_data[client][15]), time = 150),
                        "Final Fitout (1 Month)" : time_check(time_from_present(joint_data[client][25]), time=30),
                        "Financing Completed" : finance,
                        "Project Start Date" : joint_data[client][1],
                        "Projected Opening" : joint_data[client][39],
                        "Total Months in Process" : days_to_months(time_from_present(joint_data[client][1])),
                        "Projected Total Project" : days_to_months(date_subtraction(joint_data[client][1], joint_data[client][39])),
                        "Notes" : "",
                        "Sorter" : 0
                        
                        
                        
                        
                     })
                # perform the average calculations here
                AVG_COUNT_CP += 1  
                
                            
                AVG_SUM_CP1 += int(days_to_months(time_from_present(joint_data[client][1])))
                
                AVG_SUM_CP2 += int(days_to_months(date_subtraction(joint_data[client][1], joint_data[client][39])))
            
            
            elif joint_data[client][35] == "Active Interior Construction":
            
                contruction_pipeline.append(
                    {   "Franchise ID" : client,
                        "Project Status" : joint_data[client][35],
                        "City, State" : joint_data[client][37] + ", " + joint_data[client][38],
                        "Architecturals (2.5 Months)" : time_check(date_subtraction(joint_data[client][9], joint_data[client][10])),
                        "Permitting (2.5 Months)" : time_check(date_subtraction(joint_data[client][10], joint_data[client][11])),
                        "Active Construction (5 months)" : time_check(time_from_present(joint_data[client][15]), time = 150),
                        "Final Fitout (1 Month)" : time_check(time_from_present(joint_data[client][25]), time=30),
                        "Financing Completed" : finance,
                        "Project Start Date" : joint_data[client][1],
                        "Projected Opening" : joint_data[client][39],
                        "Total Months in Process" : days_to_months(time_from_present(joint_data[client][1])),
                        "Projected Total Project" : days_to_months(date_subtraction(joint_data[client][1], joint_data[client][39])),
                        "Notes" : "",
                        "Sorter" : 1
                        
                        
                        
                        
                     })
                
                # perform the average calculations here
                AVG_COUNT_CP += 1  
                
                            
                AVG_SUM_CP1 += int(days_to_months(time_from_present(joint_data[client][1])))
                
                AVG_SUM_CP2 += int(days_to_months(date_subtraction(joint_data[client][1], joint_data[client][39])))
                
                
                
            elif joint_data[client][35] == "Pre-Construction":
                
                contruction_pipeline.append(
                    {   "Franchise ID" : client,
                        "Project Status" : joint_data[client][35],
                        "City, State" : joint_data[client][37] + ", " + joint_data[client][38],
                        "Architecturals (2.5 Months)" : time_check(date_subtraction(joint_data[client][9], joint_data[client][10])),
                        "Permitting (2.5 Months)" : time_check(date_subtraction(joint_data[client][10], joint_data[client][11])),
                        "Active Construction (5 months)" : "PRE-CONSTRUCTION",
                        "Final Fitout (1 Month)" : time_check(time_from_present(joint_data[client][25]), time=30),
                        "Financing Completed" : finance,
                        "Project Start Date" : joint_data[client][1],
                        "Projected Opening" : joint_data[client][39],
                        "Total Months in Process" : days_to_months(time_from_present(joint_data[client][1])),
                        "Projected Total Project" : days_to_months(date_subtraction(joint_data[client][1], joint_data[client][39])),
                        "Notes" : "",
                        "Sorter" : 2
                        
                        
                        
                     })
                
                # perform the average calculations here
                AVG_COUNT_CP += 1  
                
                            
                AVG_SUM_CP1 += int(days_to_months(time_from_present(joint_data[client][1])))
                
                AVG_SUM_CP2 += int(days_to_months(date_subtraction(joint_data[client][1], joint_data[client][39])))
                
                
                
                
                
            elif joint_data[client][35] == "Out For Building Permit":
                
                contruction_pipeline.append(
                    {   "Franchise ID" : client,
                        "Project Status" : joint_data[client][35],
                        "City, State" : joint_data[client][37] + ", " + joint_data[client][38],
                        "Architecturals (2.5 Months)" : time_check(date_subtraction(joint_data[client][9], joint_data[client][10]), time = 75),
                        "Permitting (2.5 Months)" : time_check(time_from_present(joint_data[client][10]), time = 75),
                        "Active Construction (5 months)" : time_check(date_subtraction(joint_data[client][15], joint_data[client][25])),
                        "Final Fitout (1 Month)" : time_check(time_from_present(joint_data[client][25]), time=30),
                        "Financing Completed" : finance,
                        "Project Start Date" : joint_data[client][1],
                        "Projected Opening" : joint_data[client][39],
                        "Total Months in Process" : days_to_months(time_from_present(joint_data[client][1])),
                        "Projected Total Project" : days_to_months(date_subtraction(joint_data[client][1], joint_data[client][39])),
                        "Notes" : "",
                        "Sorter" : 3
                        
                        
                        
                     })
                
                # perform the average calculations here
                AVG_COUNT_CP += 1  
                
                            
                AVG_SUM_CP1 += int(days_to_months(time_from_present(joint_data[client][1])))
                
                AVG_SUM_CP2 += int(days_to_months(date_subtraction(joint_data[client][1], joint_data[client][39])))
                
                
                
                
                
            elif joint_data[client][35] ==  "Architectural Design":
                
                contruction_pipeline.append(
                    {   "Franchise ID" : client,
                        "Project Status" : joint_data[client][35],
                        "City, State" : joint_data[client][37] + ", " + joint_data[client][38],
                        "Architecturals (2.5 Months)" :time_check(time_from_present(joint_data[client][9]), time=75),
                        "Permitting (2.5 Months)" : time_check(date_subtraction(joint_data[client][10], joint_data[client][11])),
                        "Active Construction (5 months)" : time_check(date_subtraction(joint_data[client][15], joint_data[client][25])),
                        "Final Fitout (1 Month)" : time_check(time_from_present(joint_data[client][25]), time=30),
                        "Financing Completed" : finance,
                        "Project Start Date" : joint_data[client][1],
                        "Projected Opening" : joint_data[client][39],
                        "Total Months in Process" : days_to_months(time_from_present(joint_data[client][1])),
                        "Projected Total Project" : days_to_months(date_subtraction(joint_data[client][1], joint_data[client][39])),
                        "Notes" : "",
                        "Sorter" : 4
                        
                        
                        
                     })
                
                # perform the average calculations here
                AVG_COUNT_CP += 1  
                
                            
                AVG_SUM_CP1 += int(days_to_months(time_from_present(joint_data[client][1])))
                
                AVG_SUM_CP2 += int(days_to_months(date_subtraction(joint_data[client][1], joint_data[client][39])))
                
                
                
                
                
            elif joint_data[client][35] == "Ground Up Architecturals":
                
                contruction_pipeline.append(
                    {   "Franchise ID" : client,
                        "Project Status" : joint_data[client][35],
                        "City, State" : joint_data[client][37] + ", " + joint_data[client][38],
                        "Architecturals (2.5 Months)" : time_check(time_from_present(joint_data[client][9]), time=75),
                        "Permitting (2.5 Months)" : time_check(date_subtraction(joint_data[client][10], joint_data[client][11])),
                        "Active Construction (5 months)" : time_check(date_subtraction(joint_data[client][15], joint_data[client][25])),
                        "Final Fitout (1 Month)" : time_check(time_from_present(joint_data[client][25]), time=30),
                        "Financing Completed" : finance,
                        "Project Start Date" : joint_data[client][1],
                        "Projected Opening" : joint_data[client][39],
                        "Total Months in Process" : days_to_months(time_from_present(joint_data[client][1])),
                        "Projected Total Project" : days_to_months(date_subtraction(joint_data[client][1], joint_data[client][39])),
                        "Notes" : "",
                        "Sorter" : 5
                        
                      
                        
                     })
                # perform the average calculations here
                AVG_COUNT_CP += 1  
                
                            
                AVG_SUM_CP1 += int(days_to_months(time_from_present(joint_data[client][1])))
                AVG_SUM_CP2 += int(days_to_months(date_subtraction(joint_data[client][1], joint_data[client][39])))
                
                
                
            
            
            

            
# ORDER CONTRUCTION_PIPELINE LIST BY STATUS HERE:
construction_sorted = []
construction_sorted = sorted(contruction_pipeline, key=itemgetter('Sorter'))
    
# delete sorter
for client in construction_sorted:
    del client['Sorter']
            
            
            
            

realestate_pipeline = []


AVG_SUM_RE1 = 0
AVG_SUM_RE2 = 0


AVG_COUNT_RE = 0


for c, client in enumerate(joint_data):
    if c == 0:
        pass
    
    else:
        if joint_data[client][1].strip() == '--':
            pass
        else:
      
            # FILTERING BY STATUS
            if joint_data[client][35] == "Lease Negotiations (LOI Signed)":
            
                realestate_pipeline.append({  
                        "Franchise ID" : client,
                        "Project Status" : joint_data[client][35],
                        "City" : joint_data[client][37],
                        "State/Province" : joint_data[client][38],
                        "Site Selection (45 days)" : time_check(date_subtraction(joint_data[client][2], joint_data[client][3]), time = 45),
                        "LOI Negotiations (5 months)" : time_check(date_subtraction(joint_data[client][4], joint_data[client][6]), time = 150),
                        "Lease Negotiations (3 months)" : time_check(time_from_present(joint_data[client][6]), time=90),
                        "Expected Opening Date" : joint_data[client][39],
                        "Months in Process" : days_to_months(time_from_present(joint_data[client][1])),
                        "Projected Total Months" : days_to_months(date_subtraction(joint_data[client][1], joint_data[client][39])),
                        "Sorter" : 0
                     })
                
                # PERFORM AVERAGE CALCS HERE
                AVG_COUNT_RE += 1
                
                AVG_SUM_RE1 += int(days_to_months(time_from_present(joint_data[client][1])))
                AVG_SUM_RE2 += int(days_to_months(date_subtraction(joint_data[client][1], joint_data[client][39])))
                
                
              
                
            elif joint_data[client][35] == "LOI Negotiations":
                
                realestate_pipeline.append({  
                        "Franchise ID" : client,
                        "Project Status" : joint_data[client][35],
                        "City" : joint_data[client][37],
                        "State/Province" : joint_data[client][38],
                        "Site Selection (45 days)" : time_check(date_subtraction(joint_data[client][2], joint_data[client][3]), time = 45),
                        "LOI Negotiations (5 months)" : time_check(time_from_present(joint_data[client][4]), time=150),
                        "Lease Negotiations (3 months)" : time_check(date_subtraction(joint_data[client][6], joint_data[client][9]), time = 90),
                        "Expected Opening Date" : joint_data[client][39],
                        "Months in Process" : days_to_months(time_from_present(joint_data[client][1])),
                        "Projected Total Months" : days_to_months(date_subtraction(joint_data[client][1], joint_data[client][39])),
                        "Sorter" : 1
                     })
                
                # PERFORM AVERAGE CALCS HERE
                AVG_COUNT_RE += 1
                
                AVG_SUM_RE1 += int(days_to_months(time_from_present(joint_data[client][1])))
                AVG_SUM_RE2 += int(days_to_months(date_subtraction(joint_data[client][1], joint_data[client][39])))
                
                
                
            elif joint_data[client][35] == "Site Selection":
                
                realestate_pipeline.append({  
                        "Franchise ID" : client,
                        "Project Status" : joint_data[client][35],
                        "City" : joint_data[client][37],
                        "State/Province" : joint_data[client][38],
                        "Site Selection (45 days)" : time_check(time_from_present(joint_data[client][2]), time = 45),
                        "LOI Negotiations (5 months)" : time_check(date_subtraction(joint_data[client][4], joint_data[client][6]), time = 150),
                        "Lease Negotiations (3 months)" : time_check(date_subtraction(joint_data[client][6],joint_data[client][9]), time = 90),
                        "Expected Opening Date" : joint_data[client][39],
                        "Months in Process" : days_to_months(time_from_present(joint_data[client][1])),
                        "Projected Total Months" : days_to_months(date_subtraction(joint_data[client][1], joint_data[client][39])),
                        "Sorter" : 2
                     })
                
                # PERFORM AVERAGE CALCS HERE
                AVG_COUNT_RE += 1
                
                AVG_SUM_RE1 += int(days_to_months(time_from_present(joint_data[client][1])))
                AVG_SUM_RE2 += int(days_to_months(date_subtraction(joint_data[client][1], joint_data[client][39])))
                
                
           
        
                
# ORDER REALESTATE_PIPELINE BY STATUS HERE:
    
    
real_estate_sorted = sorted(realestate_pipeline, key=itemgetter('Sorter'))
    
# delete sorter
for client in real_estate_sorted:
    del client['Sorter']
    
    
    
AVERAGE_SUM = 0
AVERAGE_COUNT = 0

open_schools = []
largest = None
for c, client in enumerate(joint_data):
    if c == 0:
        pass
    
    elif joint_data[client][28] != "--":
            open_schools.append({  
                     "Franchise ID" : client,
                     "City, State" : joint_data[client][37] + ", " + joint_data[client][38],
                     "Open Date" : month_year_presenter(joint_data[client][28]),
                     "Months to Open" : time_check(date_subtraction(joint_data[client][1],joint_data[client][28]), time = 0),
                     "DaysInt" : days_int(time_from_present(joint_data[client][28]))
                 })
            AVERAGE_SUM += days_int(date_subtraction(joint_data[client][1],joint_data[client][28]))
            AVERAGE_COUNT += 1
            
        
        
#   ORDER OPEN_SCHOOLS BY OPEN DATE HERE 

open_schools_sorted = sorted(open_schools, key=itemgetter('DaysInt'))


# ORDER JOINT_DATA BY CUSTOMER ID

customer_info = {}

for i in sorted(joint_data):
    customer_info[i] = joint_data[i]
    

# REAL ESTATE PIPELINE TABLE ITERATION


wb = Workbook()
ws = wb.active
ws.title = "Real Estate Pipeline"

titles = list(real_estate_sorted[0])
values = []

ws.append(titles)
row = 1
for client in real_estate_sorted:
    column = 0
    row += 1
    for key in client:
        column += 1
        char = get_column_letter(column)
        
        #Conditional coloring of the DEV PROCESS
        if 5 <= column <= 7:
            if "Off" in client[key]:
                ws[char + str(row)].fill = PatternFill(fill_type = "solid",
                                                       start_color = "d78f9a",
                                                       end_color = "d78f9a")
            elif "On" in client[key]:
                ws[char + str(row)].fill = PatternFill(fill_type = "solid",
                                                       start_color = "7fe085",
                                                       end_color = "7fe085")
                
                
        ws[char + str(row)].border = Border(left = Side(border_style = "medium", color = 'b2b2b2'),
                                            right = Side(border_style = "medium", color = 'b2b2b2'),
                                            top = Side(border_style = "medium", color = 'b2b2b2'),
                                            bottom = Side(border_style = "medium", color = 'b2b2b2'))
                                            
        ws[char + str(row)].value = client[key]
        
        

        

# CONTRUCTION PIPELINE TABLE ITERATION

wb.create_sheet("Construction Pipeline")
ws = wb["Construction Pipeline"]

titles = list(construction_sorted[0])
values = []

ws.append(titles)
row = 1
for client in construction_sorted:
    column = 0
    row += 1
    for key in client:
        column += 1
        char = get_column_letter(column)
        
        #Conditional coloring of the DEV PROCESS
        if 4 <= column <= 7:
            if "Off" in client[key]:
                ws[char + str(row)].fill = PatternFill(fill_type = "solid",
                                                       start_color = "d78f9a",
                                                       end_color = "d78f9a")
            elif "On" in client[key]:
                ws[char + str(row)].fill = PatternFill(fill_type = "solid",
                                                       start_color = "7fe085",
                                                       end_color = "7fe085")
            elif "PRE-CONSTRUCTION" in client[key]:
                ws[char + str(row)].fill = PatternFill(fill_type = "solid",
                                                       start_color = "FFA400",
                                                       end_color = "FFA400")
                
        ws[char + str(row)].border = Border(left = Side(border_style = "medium", color = 'b2b2b2'),
                                            right = Side(border_style = "medium", color = 'b2b2b2'),
                                            top = Side(border_style = "medium", color = 'b2b2b2'),
                                            bottom = Side(border_style = "medium", color = 'b2b2b2'))
                                            
        ws[char + str(row)].value = client[key]
        
        
        
# OPEN SCHOOLS TABLE ITERATION
wb.create_sheet("Open Schools")
ws = wb["Open Schools"]


# delete days_int
for client in open_schools_sorted:
    del client['DaysInt']


titles = list(open_schools[0])
values = []

ws.append(titles)
row = 1
for client in open_schools_sorted:
    column = 0
    row += 1
    for key in client:
        column += 1
        char = get_column_letter(column)
        ws[char + str(row)].value = client[key]
        if column == 4:
            if row > len(open_schools_sorted):
                ws[get_column_letter(column-1) + str(row+1)].value = "Average time to Open:"
                ws[char + str(row+1)].value = time_check(pd.to_timedelta(AVERAGE_SUM/AVERAGE_COUNT, unit='D') , time=0)
                ws[get_column_letter(column-1) + str(row+1)].fill = PatternFill(fill_type = 'solid',
                                                         start_color = '99ccff',
                                                         end_color = '99ccff')
                ws[char + str(row+1)].fill = PatternFill(fill_type = 'solid',
                                                         start_color = '99ccff',
                                                         end_color = '99ccff')



# CREATE CUSTOMER DATABASE SHEET

wb.create_sheet("Client List")
ws = wb["Client List"]

client_list = []

# loop through the dictionary, and during each iteration, pull and append the data you need to the list.

for c, client in enumerate(customer_info):
    client_list.append(
        {   "Franchise ID" : client,
            "Project Status" : joint_data[client][35],
            "City, State" : joint_data[client][37] + ", " + joint_data[client][38],
            "Project Start Date" : joint_data[client][32],
            "Total Months in Process" : days_to_months(time_from_present(joint_data[client][39])),
            "Notes" : ""
         })

client_list.pop()

titles = list(client_list[0])
ws.append(titles)
row = 1
for client in client_list:
    column = 0
    row += 1
    for key in client:
        column += 1
        char = get_column_letter(column)
        ws[char + str(row)].value = client[key]

# STYLE
# Loop through all the cells with data;

count_row = 0
for row in ws:
    if not all([cell.value is None for cell in row]):
        count_row += 1

count_column = 0
for column in ws.iter_cols():
    if not all([cell.value is None for cell in column]):
        count_column += 1
        
for row in range(1, count_row+1):
    for col in range(1, count_column+1):
        
        col_letter = get_column_letter(col)
        
        if row == 1:
            ws[col_letter + str(row)].font = Font(bold = True)
            ws[col_letter + str(row)].fill = PatternFill(fill_type = 'solid',
                                                         start_color = 'BAB7B5',
                                                         end_color = 'BAB7B5'
                                                         )
        col_letter = get_column_letter(col)
        ws.column_dimensions[col_letter].width = 30
        ws.row_dimensions[row].height = 20



# open real estate pipeline WS

ws = wb["Real Estate Pipeline"]

count_row = 0
for row in ws:
    if not all([cell.value is None for cell in row]):
        count_row += 1

count_column = 1
for column in ws.iter_cols():
    if not all([cell.value is None for cell in column]):
        count_column += 1
        
for row in range(1, count_row+1):
    for col in range(1, count_column):
        #if 5 <= count_column <= 7:
            
        
        col_letter = get_column_letter(col)
        
        if row == 1:
            ws[col_letter + str(row)].font = Font(bold = True)
            ws[col_letter + str(row)].fill = PatternFill(fill_type = 'solid',
                                                         start_color = 'BAB7B5',
                                                         end_color = 'BAB7B5'
                                                         )
        col_letter = get_column_letter(col+1)
        
        ws.column_dimensions[col_letter].width = 30
        ws.row_dimensions[row].height = 20

ws.move_range("A1:J" + str(count_row+3), rows=2, cols=1)
ws.column_dimensions["A"].width = 7


# TWEAKS


DATE = str(datetime.datetime.now().strftime('%m-%d-%Y'))

ws.merge_cells("B2:C2")

ws["B2"].value = "Report Generated on " + DATE


AVG_VALUE_RE1 = AVG_SUM_RE1 / AVG_COUNT_RE

AVG_VALUE_RE2 = AVG_SUM_RE2 / AVG_COUNT_RE


count_row = 0
for row in ws:
    if not all([cell.value is None for cell in row]):
        count_row += 1
        
ws["J"+ str(count_row+2)].value = "AVERAGE = " + f'{AVG_VALUE_RE1:.2f}'
ws["K"+ str(count_row+2)].value = "AVERAGE = " + f'{AVG_VALUE_RE2:.2f}'

set_border(ws, "J"+ str(count_row+2) + ":K"+ str(count_row+2))

ws.merge_cells("F2:H2")
ws["f2"].value = "Real Estate Pipeline"
ws['f2'].alignment = Alignment(horizontal = "center", vertical = "center")

ws["F2"].font = Font(bold = True)
ws["F2"].fill = PatternFill(fill_type = 'solid',
                                             start_color = 'BAB7B5',
                                             end_color = 'BAB7B5'
                                             )

set_border(ws, "f2:H"+ str(count_row + 1))

for row in range(1, 100):
    
    ws.row_dimensions[row].height = 30
    
    for column in range(1,100):
        char = get_column_letter(column)
        
        if 5 < column < 9:
            ws.column_dimensions[char].width = 37
        
        ws[char + str(row)].alignment = Alignment(horizontal = "center", vertical = "center")
        

        
# open construction pipeline WS

ws = wb["Construction Pipeline"]


count_row = 0
for row in ws:
    if not all([cell.value is None for cell in row]):
        count_row += 1

count_column = 0
for column in ws.iter_cols():
    if not all([cell.value is None for cell in column]):
        count_column += 1
        
for row in range(1, count_row+1):
    for col in range(1, count_column+1):
        
        col_letter = get_column_letter(col)
        
        if row == 1:
            ws[col_letter + str(row)].font = Font(bold = True)
            ws[col_letter + str(row)].fill = PatternFill(fill_type = 'solid',
                                                         start_color = 'BAB7B5',
                                                         end_color = 'BAB7B5'
                                                         )
        col_letter = get_column_letter(col+1)
        ws.column_dimensions[col_letter].width = 30
        ws.row_dimensions[row].height = 20
        
ws.move_range("A1:M" + str(count_row+3), rows=2, cols=1)
ws.column_dimensions["A"].width = 7


#TWEAKS

AVG_VALUE_CP1 = AVG_SUM_CP1 / AVG_COUNT_CP

AVG_VALUE_CP2 = AVG_SUM_CP2 / AVG_COUNT_CP


ws.merge_cells("B2:C2")

ws["B2"].value = "Report Generated on " + DATE


count_row = 0
for row in ws:
    if not all([cell.value is None for cell in row]):
        count_row += 1
        
ws["L"+ str(count_row+2)].value = "AVERAGE = " + f'{AVG_VALUE_CP1:.2f}'
ws["M"+ str(count_row+2)].value = "AVERAGE = " + f'{AVG_VALUE_CP2:.2f}'

set_border(ws, "L"+ str(count_row+2) + ":M"+ str(count_row+2))

ws.merge_cells("E2:H2")
ws["e2"].value = "Development Journey"
ws['e2'].alignment = Alignment(horizontal = "center", vertical = "center")

ws["e2"].font = Font(bold = True)
ws["e2"].fill = PatternFill(fill_type = 'solid',
                                             start_color = 'BAB7B5',
                                             end_color = 'BAB7B5'
                                             )

set_border(ws, "E2:H" + str(count_row+1))

for row in range(1, 100):
    
    ws.row_dimensions[row].height = 30
    
    for column in range(1,100):
        char = get_column_letter(column)
        
        if 4 < column < 9:
            ws.column_dimensions[char].width = 37
        
        ws[char + str(row)].alignment = Alignment(horizontal = "center", vertical = "center")
        
        


# open schools formatting

ws = wb["Open Schools"]


count_row = 0
for row in ws:
    if not all([cell.value is None for cell in row]):
        count_row += 1

count_column = 0
for column in ws.iter_cols():
    if not all([cell.value is None for cell in column]):
        count_column += 1
        
for row in range(1, count_row+1):
    for col in range(1, count_column+1):
        
        col_letter = get_column_letter(col)
        
        if row == 1:
            ws[col_letter + str(row)].font = Font(bold = True)
            ws[col_letter + str(row)].fill = PatternFill(fill_type = 'solid',
                                                         start_color = 'BAB7B5',
                                                         end_color = 'BAB7B5')
        col_letter = get_column_letter(col+1)    
        ws.column_dimensions[col_letter].width = 30
        ws.row_dimensions[row].height = 20

ws.move_range("A1:D" + str(count_row+3), rows=2, cols=1)
ws.column_dimensions["A"].width = 7



for row in range(1, 100):
    
    ws.row_dimensions[row].height = 30
    
    for column in range(1,100):
        char = get_column_letter(column)
        
        ws[char + str(row)].alignment = Alignment(horizontal = "center", vertical = "center")
        
        
for works in wb.worksheets:
    works.sheet_view.zoomScale = 80


 


wb.save('result.xlsx') 




































    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    